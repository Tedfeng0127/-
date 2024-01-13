import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
import re
import datetime
import time




def get_total_page_number(url):
    """
    傳入每個法規類別項目的頁面url，
    回傳每個法規類別項目的頁面總數
    """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    page_info = soup.find("li", class_="pageinfo")
    # 如果該項目沒有法規或法規筆數小於等於10筆，不會有頁面數量資訊，回傳0
    if not page_info:
        return 0
    page_info = re.sub("\s", "", page_info.text)
    search1 = re.search("頁次：", page_info)
    search2 = re.search("，每頁顯示", page_info)
    start, end = search1.span()[1], search2.span()[0]
    total_page = page_info[start:end].split("/")[-1]
    return int(total_page)


def get_law_title_and_info(url):
    """
    傳入頁面網址，取得頁面上法條的相關資訊，
    包括發布(更新)日期、標題、連結、法規類別
    """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table", class_="table table-hover tab-list tab-category")
    # 如果沒有table，代表此類別沒有法條，直接return
    if not table:
        return []
    table_row = table.find_all("tr")
    return_list = []
    for row in table_row:
        row_data = row.find_all("td")
        date = row_data[1].text
        date_split = date.split(".")
        date_split[0] = str(int(date_split[0]))
        date = ".".join(date_split)
        title = row_data[2]
        if title.find("span", class_="label-fei") and title.find("span", class_="label-fei").text == ("廢" or "廢/停"):
            continue
        title = title.find("a").text
        title_link = "https://oaout.epa.gov.tw/law/" + row_data[2].find("a").get("href")
        law_category = row_data[3].text
        if law_category in ["法律", "命令"]:
            return_list.append((date, title, title_link, law_category))
    return return_list


def get_law_release_date_worksheet1(url):
    """
    傳入第一個worksheet的法條(環保署網站上的法條)的頁面連結，取得該法條的相關資訊
    """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table", class_="table table-bordered tab-edit")
    table_row = table.find_all("tr")
    th_td_dict = {tr.find("th").text.replace("：", ""):tr.find("td").text.replace(" ", "").replace("\r", "").replace("\n", "")
                  for tr in table_row}
    if "修正日期" in th_td_dict:
        date = th_td_dict["修正日期"]
        date_type = "修正日期"
    elif "公發布日" in th_td_dict:
        date = th_td_dict["公發布日"]
        date_type = "公發布日"
    else:
        return None
    year = date[date.index("國") + 1:date.index("年")]
    month = date[date.index("年") + 1:date.index("月")]
    day = date[date.index("月") + 1:date.index("日")]
    return (date_type, f"{year}.{month}.{day}")


def get_law_release_date_worksheet2(url):
    """
    傳入第二個worksheet內的法條的頁面連結，取得該法條的發布日期
    (第二個worlsheet內的法條是環保署法條之外的那些法條)
    """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table", class_="table")
    th_td_dict = {tr.find("th").text.replace("：", ""):tr.find("td").text.replace(" ", "").replace("\r", "").replace("\n", "")
                  for tr in table.find_all("tr")}
    if "修正日期" in th_td_dict:
        date = th_td_dict["修正日期"]
    elif "公發布日" in th_td_dict:
        date = th_td_dict["公發布日"]
    else:
        return None
    year = date[date.index("國") + 1:date.index("年")]
    month = date[date.index("年") + 1:date.index("月")]
    day = date[date.index("月") + 1:date.index("日")]
    return f"{year}.{month}.{day}"


def check_law_update_date_and_update_excel_file_worksheet1(file_name):
    """
    傳入要檢查的excel檔檔名，檢查excel檔中第一個worksheet的法條(環保署網站上的法條)的最後修正日期有沒有更新，
    並在excel檔新增一個column，寫入最新的最後修正日期
    """
    # excel檔中第一個worksheet的row、column數
    workbook = openpyxl.load_workbook(f"{file_name}.xlsx")
    sheet = workbook.worksheets[0]
    row_num = sheet.max_row
    col_num = sheet.max_column
    new_col_num = col_num + 1
    print(f"此worksheet中目前有{row_num}個row和{col_num}個column的資料")
    
    # 寫入最新的最後修正日期column
    date_today = datetime.date.today()
    font = Font(name="Calibri", size=11)
    sheet.cell(row=1, column=new_col_num).value = date_today
    sheet.cell(row=1, column=new_col_num).number_format = "yyyymmdd"
    sheet.cell(row=1, column=new_col_num).font = font
    sheet.cell(row=2, column=new_col_num).value = "最後修正日期"
    sheet.cell(row=2, column=new_col_num).font = font
    column = sheet["A"]
    for cell in column:
        if not cell.hyperlink:
            pass
        else:
            r = cell.row
            curr_cell = sheet.cell(row=r, column=new_col_num)
            law_url = cell.hyperlink.target
            date_of_law = sheet.cell(row=r, column=col_num).value
            time.sleep(1)
            date_type, date = get_law_release_date_worksheet1(law_url)
            font_red = Font(color="00FF0000", name="Calibri", size=11)    # 紅色字體
            font_blue = Font(color="000000FF", name="Calibri", size=11)   # 藍色字體
            curr_cell.value = date
            curr_cell.number_format = "yyy.mm.dd"
            # 最後修正日期：有更新 -> 紅字  /  沒更新 -> 藍字
            curr_cell.font = font_red if date_of_law != date else font_blue
            print(cell.value)
    file_name_date = date_today.strftime("%Y%m%d")
    workbook.save(f"QF9007002 EMS Compliance Obligation Review Form {file_name_date}.xlsx")


def check_law_update_date_and_update_excel_file_worksheet2(file_name):
    """
    傳入要檢查的excel檔的檔名，檢查excel檔中第二個worksheet的法條的最後修正日期有沒有更新，
    並在excel檔新增一個column，寫入最新的最後修正日期
    """
    # excel檔中第二個worksheet的相關資訊
    workbook = openpyxl.load_workbook(f"{file_name}.xlsx")
    sheet = workbook.worksheets[1]
    row_num = sheet.max_row
    col_num = sheet.max_column
    new_col_num = col_num + 1
    print(f"此worksheet中目前有{row_num}個row和{col_num}個column的資料")

    date_today = datetime.date.today()
    font = Font(name="Calibri", size=11)
    sheet.cell(row=1, column=new_col_num).value = date_today
    sheet.cell(row=1, column=new_col_num).number_format = "yyyymmdd"
    sheet.cell(row=1, column=new_col_num).font = font
    sheet.cell(row=2, column=new_col_num).value = "最後修正日期"
    sheet.cell(row=2, column=new_col_num).font = font
    column = sheet["A"]
    for cell in column:
        if not cell.hyperlink:
            pass
        else:
            r = cell.row
            curr_cell = sheet.cell(row=r, column=new_col_num)
            law_url = cell.hyperlink.target
            date_of_law = sheet.cell(row=r, column=col_num).value
            time.sleep(1)
            date = get_law_release_date_worksheet2(law_url)
            font_red = Font(color="00FF0000", name="Calibri", size=11)
            font_blue = Font(color="000000FF", name="Calibri", size=11)
            curr_cell.value = date
            curr_cell.number_format = "yyy.mm.dd"
            # 最後修正日期：更新 -> 紅字  /  無更新 -> 藍字
            curr_cell.font = font_red if date_of_law != date else font_blue
            print(cell.value)
    workbook.save(f"{file_name}.xlsx")
    
    
def get_all_law_in_excel_sheet(file_name):
    """
    傳入excel檔的檔名，讀取檔案的第一個欄位(法條類別和法條的欄位)，
    回傳以法條類別為key，法條為value的dict
    """
    workbook = openpyxl.load_workbook(f"{file_name}.xlsx")
    sheet = workbook.worksheets[0]
    column = sheet["A"]
    
    return_dict = {}
    # traverse整個column
    for cell in column:
        if cell.row <= 2:
            pass
        else:
            cell_val = cell.value
            if cell_val[0].isnumeric():
                # key是法條類別的編號和類別的名稱
                key = (int(cell_val.split(". ")[0]), cell_val.split(". ")[1])
                return_dict[key] = []
                continue
            # value是法條在該類別的編號和法條名稱
            value = (int(cell_val.split(")")[0].replace("(", "")) , cell_val.split(")")[1])
            return_dict[key].append(value)
    workbook.close()
    return return_dict


def get_all_law_on_regulation_system():
    """
    取得行政院環境保護署主管法規查詢系統上所有需要的法條類別和法條內容，
    頁面網址：https://oaout.epa.gov.tw/law/LawCategoryMain.aspx
    """
    # 取得環保署法規系統的所有法條(每個類別的每個項目的每個法條)
    url = "https://oaout.epa.gov.tw/law/LawCategoryMain.aspx"
    response = requests.get(url)
    s1 = BeautifulSoup(response.text, "html.parser")
    s2 = s1.find("ul", id="ctl00_cp_content_ulLawCategory")
    s3 = s2.find("li")
    all_category = [s3] + list(s3.find_next_siblings("li"))
    category_dict = {}
    for c in all_category:
        category = c.find("a").text.replace("\n", "").replace("\r", "").replace("已展開", "").replace("已收合", "")
        items = c.find("ul").find_all("li")
        category_data = []
        for item in items:
            item_count = item.find("span").text
            item_name = item.text.replace(item_count, "")
            item_url = f"https://oaout.epa.gov.tw{item.find('a').get('href')}"
            category_data.append([item_name, int(item_count), item_url])
        category_dict[category] = category_data
    
    # 檢查需要哪些項目的法條，只保留需要的法條
    item_list = [
    "環境綜合管理", "環境教育", "環境影響評估",
    "空氣品質管理", "固定源空污防制", "移動源空污防制", "室內空氣品質管理", "噪音與電磁波管理",
    "水污染防治", "飲用水管理",
    "廢棄物清理", "資源回收再利用",
    "溫室氣體管理",
    "環境綜合計畫", "公害糾紛處理",
    "環境衛生",
    "應回收廢棄物",
    "毒化物管理", "環境用藥管理",
    "環保人員訓練"
    ]
    # 存法條資料的dict
    return_dict = {}
    categories = list(category_dict.keys())
    # traverse每個category
    for category in categories:
        items = category_dict[category]
        # traverse每個item
        for item in items:
            item_name = item[0]
            item_count = item[1]
            item_url = item[2]
            # 判斷item是不是需要檢查的item
            if item_name in item_list:
                total_page_number = get_total_page_number(item_url)
                
                # 這個item內的所有法條
                law_infos = []
                # item內沒有法條
                if item_count == 0:
                    continue
                # item內有法條，但法條數目不到10筆，可以全部顯示在一個頁面不用換頁
                elif total_page_number == 0:
                    law_info = get_law_title_and_info(item_url)
                    if len(law_info) == 0:
                        pass
                    else:
                        for law in law_info:
                            law_date = law[0]
                            law_title = law[1]
                            law_url = law[2]
                            law_type = law[3]
                            law_infos.append((law_date, law_type, law_title, law_url))
                # item內有法條，且法條數目超過10筆，需要換頁才能顯示所有法條
                else:
                    for page in range(1, total_page_number + 1):
                        url = item_url + f"&page={page}"
                        law_info = get_law_title_and_info(url)
                        if len(law_info) == 0:
                            pass
                        else:
                            for law in law_info:
                                law_date = law[0]
                                law_title = law[1]
                                law_url = law[2]
                                law_type = law[3]
                                law_infos.append((law_date, law_type, law_title, law_url))
                return_dict[item_name] = law_infos
    return return_dict


def get_item_number_and_law_number_of_new_law(law_in_excel, law_on_web):
    """
    傳入excel檔中的法條資料和從環保署網站上抓下來的法條資料，
    回傳一個dict，key是新法條的類別編號和法條編號，value是新法條的名稱、日期、類型(命令or法律)、連結，
    新法條是從環保署網站上抓到，但還沒存進excel檔中的法條
    """
    # 儲存新法條在excel檔案中的項目編號和法條編號
    new_law_list = []
    # traverse excel檔案中所有法條類別
    for item_num, item in law_in_excel:
        # 環保署網站上這個類別的所有法條
        web_law = law_on_web.get(item)
        # excel檔案中這個類別的所有法條
        excel_law = [law[1] for law in law_in_excel[(item_num, item)]]
        # excel檔案中這個類別的法條數量
        law_count = len(excel_law)
        
        if not web_law:
            pass
        else:
            # traverse環保署網站上這個類別的所有法條，如果這個類別中的法條不在excel檔，
            # 代表這個法條是在上一次檢查之後才新增的，把這個法條新增到excel檔中對應的類別中
            for law in web_law:
                law_date = law[0]
                law_type = law[1]
                law_title = law[2]
                law_url = law[3]
                # 網站上出現不在execl檔案中的法條，把該法條在excel檔案中的項目編號和法條編號存起來
                if law_title not in excel_law:
                    law_num = law_count + 1
                    law_count += 1
                    new_law_list.append([(item_num, law_num), (law_date, law_type, law_title, law_url)])
                
    # 用dict把新法條的資訊儲存並回傳
    return_dict = {}
    for number, law_info in new_law_list:
        return_dict[number] = law_info
    return return_dict


def insert_new_law_into_excel_file(file_name, new_law_dict):
    """
    傳入excel檔名和新增法條的資訊，
    把新法條插入excel檔對應的項目中
    """
    workbook = openpyxl.load_workbook(f"{file_name}.xlsx")
    sheet = workbook.worksheets[0]
    column = sheet["A"]
    curr_row_num = 1
    max_row_num = sheet.max_row
    max_col_num = sheet.max_column
    new_col_num = max_col_num + 1
    
    # traverse整個worksheet
    while curr_row_num <= max_row_num:
        if curr_row_num <= 2:
            curr_row_num += 1
            continue
        # 當下所在的儲存格
        current_cell = sheet.cell(row=curr_row_num, column=1)
        current_cell_value = current_cell.value
        # 目前所在的儲存格是法條類別
        if current_cell_value[0].isnumeric():
            item_num = int(current_cell_value.split(". ")[0])
            item_title = current_cell_value.split(". ")[1]
            law_num = 0
        # 目前所在的儲存格是法條
        else:
            law_num = int(current_cell_value.split(")")[0].replace("(", ""))
            law_title = current_cell_value.split(")")[1]
        
        # 如果目前所在的法條類別中有新的法條，要在worksheet中insert新的row
        if (item_num, law_num + 1) in new_law_dict:
            new_law_date, new_law_type, new_law_title, new_law_url = new_law_dict.pop((item_num, law_num + 1))
            new_row_num = curr_row_num + 1
            new_law_num = law_num + 1
            # 在worksheet insert row，並寫入新法條的資訊
            sheet.insert_rows(new_row_num)
            sheet.cell(row=new_row_num, column=1).value = f"({new_law_num}){new_law_title}"
            sheet.cell(row=new_row_num, column=1).hyperlink = new_law_url
            sheet.cell(row=new_row_num, column=1).style = "Hyperlink"
            sheet.cell(row=new_row_num, column=1).font = Font(name="Calibri", size=11, color="000000FF", underline="single")
            sheet.cell(row=new_row_num, column=2).value = new_law_type
            sheet.cell(row=new_row_num, column=2).font = Font(name="Calibri", size=11, color="0033CCCC")
            sheet.cell(row=new_row_num, column=max_col_num).value = new_law_date
            sheet.cell(row=new_row_num, column=max_col_num).number_format = "yyy.mm.dd"
            font_green = openpyxl.styles.Font(color="0033CCCC", name="Calibri", size=11)   # 綠色字體
            sheet.cell(row=new_row_num, column=max_col_num).font = font_green
            # insert一個新的row，max_row要+1
            max_row_num += 1
            # 因為insert新的row，要把後面的儲存格的hyperlink的ref指向正確的儲存格
            for row in range(max_row_num, new_row_num, -1):
                c_cell = sheet.cell(row=row, column=1)
                if c_cell.hyperlink:
                    c_cell_coordinate = c_cell.coordinate
                    c_cell.hyperlink.ref = c_cell_coordinate
            print(f"新增一筆法條：{item_num}({law_num + 1}) {new_law_title}")
        curr_row_num += 1
    workbook.save(f"{file_name}.xlsx")



if __name__ == "__main__":
    print(f"【{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}】", "start")

    previous_review_date = input("請輸入上次檢查日期(yyyymmdd)：")
    file_name = f"QF9007002 EMS Compliance Obligation Review Form {previous_review_date}"

    # 檢查excel檔內第一個worksheet的法條(環保署上的法條)的最後修正日期
    check_law_update_date_and_update_excel_file_worksheet1(file_name)
    print(f"【{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}】", "finish checking update date of worksheet1")
    # 用新的檔名開啟excel檔
    date = datetime.date.today().strftime("%Y%m%d")
    file_name = f"QF9007002 EMS Compliance Obligation Review Form {date}"
    # 檢查excel檔內第二個worksheet的法條(不屬於環保署的法條)的最後修正日期
    check_law_update_date_and_update_excel_file_worksheet2(file_name)
    print(f"【{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}】", "finish checking update date of worksheet2")

    # 抓取環保署網站上需要的法條資料
    law_on_web = get_all_law_on_regulation_system()
    print(f"【{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}】", "抓取環保署網站法條資料")

    # 讀取excel檔內的法條資料
    law_in_excel = get_all_law_in_excel_sheet(file_name)
    print(f"【{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}】", "讀取excel內的法條")

    # 根據excel內已有的法條資料和新的法條資料，把新法條的編號和資訊存在一個dict裡
    new_law_dict = get_item_number_and_law_number_of_new_law(law_in_excel, law_on_web)

    # 把新法條根據編號插入excel檔對應的類別中
    insert_new_law_into_excel_file(file_name, new_law_dict)
    print(f"【{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}】", "插入新法條")

    print(f"【{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}】", "finish")